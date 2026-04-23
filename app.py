import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
import re
import tempfile
import pandas as pd
import os
from openpyxl.comments import Comment
import zipfile
import io

# Page config
st.set_page_config(page_title="Excel IP Colorizer", layout="centered")

# Title
st.title("📊 Excel IP Colorizer")
st.caption("Upload Excel file and IP list to highlight matching IPs.")

# --- File Upload ---
excel_files = st.file_uploader("Upload Excel Files", type=["xlsx", "xls", "xlsm"], accept_multiple_files=True)

ip_file = st.file_uploader("Upload IP List File", type=["txt", "csv", "xlsx"])

# --- Show Uploaded Files ---
st.divider()
st.write("### Uploaded Files")

if excel_files:
    st.success(f"Excel files: {', '.join([f.name for f in excel_files])}")
    st.write(f"Total files selected: {len(excel_files)}")

if ip_file:
    st.success(f"IP file: {ip_file.name}")
    st.write(f"Detected IP file type: {ip_file.name}")


# --- Load IP List ---
def load_ip_list(ip_path):
    if ip_path.endswith(".txt"):
        with open(ip_path, "r") as f:
            return {line.strip() for line in f if line.strip()}

    elif ip_path.endswith(".csv"):
        df = pd.read_csv(ip_path)
        return set(df.iloc[:, 0].astype(str))

    elif ip_path.endswith(".xlsx"):
        df = pd.read_excel(ip_path)
        return set(df.iloc[:, 0].astype(str))

    else:
        raise ValueError("Unsupported IP file format")


# --- Process Function ---
def process_excel(excel_path, ip_path, output_path):
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )

    ip_pattern = re.compile(
        r"(?:25[0-5]|2[0-4]\d|1\d{2}|[1-9]?\d)"
        r"(?:\.(?:25[0-5]|2[0-4]\d|1\d{2}|[1-9]?\d)){3}"
    )
    pinging_ips = load_ip_list(ip_path)

    workbook = openpyxl.load_workbook(excel_path)

    match_count = 0
    no_match_count = 0
    mixed_count = 0

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.value:
                    continue

                if isinstance(cell.value, str):
                    ips = ip_pattern.findall(cell.value)
                    if ips:
                        matched = [ip for ip in ips if ip in pinging_ips]
                        unmatched = [ip for ip in ips if ip not in pinging_ips]
                        if len(matched) == len(ips):
                            cell.fill = green_fill
                            match_count += 1

                        elif len(matched) == 0:
                            cell.fill = red_fill
                            no_match_count += 1

                        else:
                            cell.fill = yellow_fill
                            mixed_count += 1

                            comment_text = (
                                f"Matched: {','.join(matched)}\n"
                                f"Not Matched: {','.join(unmatched)}"
                            )
                            if not cell.comment:
                                cell.comment = Comment(comment_text, "IP Checker")

    workbook.save(output_path)

    return match_count, no_match_count, mixed_count


# --- Process Button ---
if st.button("🚀 Process Files", use_container_width=True):

    if not excel_files or ip_file is None:
        st.error("Please upload at least one Excel file and an IP list file.")
    else:
        st.info(f"Processing {len(excel_files)} file(s)...")

        with st.spinner("Processing..."):
            # Save IP file with correct extension
            ip_extension = os.path.splitext(ip_file.name)[1]
            with tempfile.NamedTemporaryFile(
                delete=False, suffix=ip_extension
            ) as temp_ip:
                temp_ip.write(ip_file.read())
                ip_path = temp_ip.name

            # Create temporary directory for processed files
            temp_dir = tempfile.mkdtemp()
            
            # Process each Excel file
            all_results = []
            for excel_file in excel_files:
                # Save Excel file
                with tempfile.NamedTemporaryFile(
                    delete=False, suffix=".xlsx"
                ) as temp_excel:
                    temp_excel.write(excel_file.read())
                    excel_path = temp_excel.name

                # Output file
                output_filename = f"{os.path.splitext(excel_file.name)[0]}_processed.xlsx"
                output_path = os.path.join(temp_dir, output_filename)

                # Process file
                match_count, no_match_count, mixed_count = process_excel(
                    excel_path, ip_path, output_path
                )
                
                all_results.append({
                    "filename": excel_file.name,
                    "matched": match_count,
                    "not_matched": no_match_count,
                    "mixed": mixed_count
                })

            # Create zip file with all processed files
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for filename in os.listdir(temp_dir):
                    file_path = os.path.join(temp_dir, filename)
                    zip_file.write(file_path, arcname=filename)
            
            zip_buffer.seek(0)

        # --- Results ---
        st.divider()
        st.subheader("Results")

        # Display results for each file
        for result in all_results:
            st.success(
                f"""
                **{result['filename']}**
                ✅ Matched IPs: {result['matched']}
                ❌ Not Matched IPs: {result['not_matched']}
                ⚠️ Mixed IPs: {result['mixed']}
                """
            )

        # Download button for zip file
        st.download_button(
            label="📥 Download All Processed Files (ZIP)",
            data=zip_buffer.getvalue(),
            file_name="processed_files.zip",
            mime="application/zip",
            use_container_width=True
        )
